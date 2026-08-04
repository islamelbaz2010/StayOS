import csv
import io
from typing import Any

from app.importer.schemas import ImportRowData

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


REQUIRED_FIELDS = {
    "title",
    "description",
    "city",
    "governorate",
    "latitude",
    "longitude",
    "property_type",
    "price",
}

COLUMN_ALIASES: dict[str, str] = {
    "title": "title",
    "description": "description",
    "address": "address",
    "district": "district",
    "city": "city",
    "governorate": "governorate",
    "country": "country",
    "latitude": "latitude",
    "lat": "latitude",
    "longitude": "longitude",
    "lng": "longitude",
    "long": "longitude",
    "property type": "property_type",
    "property_type": "property_type",
    "bedrooms": "bedrooms",
    "beds": "beds",
    "bathrooms": "bathrooms",
    "max guests": "max_guests",
    "max_guests": "max_guests",
    "price": "price",
    "currency": "currency",
    "amenities": "amenities",
    "host name": "host_name",
    "host_name": "host_name",
    "host phone": "host_phone",
    "host_phone": "host_phone",
    "host email": "host_email",
    "host_email": "host_email",
    "image urls": "image_urls",
    "image_urls": "image_urls",
    "images": "image_urls",
    "status": "status",
}


def _normalize_header(header: str) -> str:
    return header.strip().lower()


def _map_headers(raw_headers: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, raw in enumerate(raw_headers):
        normalized = _normalize_header(raw)
        canonical = COLUMN_ALIASES.get(normalized, normalized)
        mapping[idx] = canonical
    return mapping


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def _parse_list(value: Any) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if v]
    raw = str(value).strip()
    if raw.startswith("[") and raw.endswith("]"):
        raw = raw[1:-1]
    return [item.strip() for item in raw.split(",") if item.strip()]


def _parse_image_urls(value: Any) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if v]
    raw = str(value).strip()
    if raw.startswith("[") and raw.endswith("]"):
        raw = raw[1:-1]
    return [url.strip() for url in raw.split(",") if url.strip()]


def _row_to_import_data(
    row_number: int,
    field_map: dict[int, str],
    values: list[Any],
) -> ImportRowData:
    raw: dict[str, Any] = {}
    for idx, canonical in field_map.items():
        if idx < len(values):
            raw[canonical] = values[idx]

    return ImportRowData(
        row_number=row_number,
        title=str(raw.get("title", "")).strip(),
        description=str(raw.get("description", "")).strip(),
        address=str(raw.get("address", "")).strip() or None,
        district=str(raw.get("district", "")).strip() or None,
        city=str(raw.get("city", "")).strip(),
        governorate=str(raw.get("governorate", "")).strip(),
        country=str(raw.get("country", "Egypt")).strip() or "Egypt",
        latitude=_parse_float(raw.get("latitude")) or 0.0,
        longitude=_parse_float(raw.get("longitude")) or 0.0,
        property_type=str(raw.get("property_type", "")).strip().upper(),
        bedrooms=_parse_int(raw.get("bedrooms")) or 0,
        beds=_parse_int(raw.get("beds")) or 1,
        bathrooms=_parse_int(raw.get("bathrooms")) or 1,
        max_guests=_parse_int(raw.get("max_guests")) or 1,
        price=_parse_int(raw.get("price")) or 0,
        currency=str(raw.get("currency", "EGP")).strip().upper() or "EGP",
        amenities=_parse_list(raw.get("amenities")),
        image_urls=_parse_image_urls(raw.get("image_urls")),
        host_name=str(raw.get("host_name", "")).strip() or None,
        host_phone=str(raw.get("host_phone", "")).strip() or None,
        host_email=str(raw.get("host_email", "")).strip() or None,
        status=str(raw.get("status", "PENDING_VERIFICATION")).strip().upper() or "PENDING_VERIFICATION",
    )


def parse_csv(content: bytes) -> list[ImportRowData]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    raw_headers = next(reader, [])
    field_map = _map_headers(raw_headers)

    rows: list[ImportRowData] = []
    for i, values in enumerate(reader, start=2):
        if not any(v.strip() for v in values if isinstance(v, str)):
            continue
        rows.append(_row_to_import_data(i, field_map, values))
    return rows


def parse_xlsx(content: bytes) -> list[ImportRowData]:
    if load_workbook is None:
        raise ImportError("openpyxl is required for Excel import")

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    raw_headers = [str(cell.value or "") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    field_map = _map_headers(raw_headers)

    rows: list[ImportRowData] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row)
        if not any(v for v in values if v is not None and str(v).strip()):
            continue
        rows.append(_row_to_import_data(i, field_map, values))

    wb.close()
    return rows


def parse_file(filename: str, content: bytes) -> list[ImportRowData]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv(content)
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return parse_xlsx(content)
    raise ValueError(f"Unsupported file format: {filename}")
